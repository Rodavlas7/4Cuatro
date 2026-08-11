import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from datetime import datetime

# ==============================================================================
# RF51: GENERADOR DE PDF PARA EXPEDIENTE DE LAPTOP
# ==============================================================================
def generar_pdf_expediente(request, contexto, num_serie):
    """
    Toma un diccionario de contexto, renderiza la plantilla HTML y la convierte a PDF.
    """
    # 1. Renderizar el HTML con los datos
    html_string = render_to_string('trazabilidad/trazabilidad_laptop_pdf.html', contexto)
    
    # 2. Generar el PDF usando WeasyPrint (base_url es necesario para cargar CSS/Imágenes si las hay)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()

    # 3. Construir la respuesta HTTP
    response = HttpResponse(pdf_file, content_type='application/pdf')
    # Usamos 'inline' para que se abra en el navegador, si quieres forzar descarga usa 'attachment'
    response['Content-Disposition'] = f'inline; filename="Expediente_{num_serie}.pdf"'
    
    return response


# ==============================================================================
# RF52: GENERADOR DE EXCEL PARA INSPECCIONES DE CALIDAD
# ==============================================================================
def generar_excel_calidad(inspecciones):
    """
    Toma una lista de diccionarios (inspecciones ya filtradas) y construye un archivo Excel.
    """
    # 1. Crear el libro y la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Calidad"

    # 2. Definir y estilizar Cabeceras
    columnas = ['Num. Inspección', 'Línea', 'Fecha', 'Hora', 'Inspector', 'Dictamen', 'Observaciones']
    ws.append(columnas)

    # Estilos del encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # Ajustar ancho de columnas aproximado
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18

    # La columna de observaciones necesita más espacio
    ws.column_dimensions['G'].width = 40 

    # 3. Llenar los datos
    for i in inspecciones:
        ws.append([
            f"#{i.get('numero')}",
            i.get('linea_nombre') or "N/A",
            i.get('fecha') or "Sin fecha",
            i.get('hora') or "Sin hora",
            i.get('empleado_nombre') or "No registrado",
            i.get('resultado_nombre') or "Pendiente",
            i.get('observaciones') or ""
        ])

    # 4. Construir la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Calidad_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    # Guardar el Excel en la respuesta HTTP
    wb.save(response)
    return response



# ==============================================================================
# RF53: GENERADOR DE EXCEL PARA CONTROL DE EMBALAJE
# ==============================================================================
def generar_excel_embalaje(embalajes):
    """
    Toma una lista de diccionarios (embalajes filtrados) y construye un archivo Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control de Embalaje"

    # Cabeceras
    columnas = ['Num. Embalaje', 'Num. Serie Laptop', 'Fecha', 'Hora', 'Tipo de Empaque']
    ws.append(columnas)

    # Estilos del encabezado (Azul para diferenciarlo del de calidad)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # Llenar los datos
    for e in embalajes:
        ws.append([
            f"#{e.get('numero')}",
            e.get('laptop_num_serie') or "N/A",
            e.get('fecha') or "Sin fecha",
            e.get('hora') or "Sin hora",
            e.get('tipo_nombre') or "Estándar"
        ])

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Control_Embalaje_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response